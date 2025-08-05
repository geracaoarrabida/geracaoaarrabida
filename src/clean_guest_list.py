from openpyxl import load_workbook
import sys
import os
from io import BytesIO
import requests

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.utils import *

# === Load existing CSV (if exists) ===
if os.path.exists(GUEST_LIST_CSV_PATH):
    existing_df = pd.read_csv(GUEST_LIST_CSV_PATH)
    existing_keys = set(zip(
        existing_df["date"].astype(str),
        existing_df["complete_name"].str.strip().str.lower(),
        existing_df["email"].str.strip().str.lower()
    ))

else:
    existing_df = pd.DataFrame()
    existing_keys = set()

# === Step 1: Download Excel from Google Sheets ===
sheet_url = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_FILE_ID}/export?format=xlsx"

response = requests.get(sheet_url)
excel_bytes = BytesIO(response.content)

# === Extract hyperlinks from "Payable Order ID" column ===
wb = load_workbook(excel_bytes, data_only=True)
sheet = wb["Respostas do Formulário 1"]

header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
payable_idx = header.index("Payable Order ID") + 1
payable_links = {}
for row in sheet.iter_rows(min_row=2):
    cell = row[payable_idx - 1]
    link = cell.hyperlink.target if cell.hyperlink else None
    payable_links[cell.row] = link

# === Read with pandas and filter ===
excel_bytes.seek(0)
df = pd.read_excel(excel_bytes, sheet_name="Respostas do Formulário 1")
df = df[(df["STATUS"] == True) & (df["PAYMENT_VALIDATION"] != False)]

# === Process new rows ===
output_rows = []
today_str = date.today().isoformat()

for idx, row in df.iterrows():
    row_number = idx + 2
    submission_date = pd.to_datetime(row["Carimbo de data/hora"]).date()
    email_registration = row["Endereço de email"]
    supporting_document = row.get("Submeta aqui o comprovativo de pagamento referente ao total de bilhetes")

    if pd.isna(supporting_document) or not str(supporting_document).strip():
        supporting_document = payable_links.get(row_number)

    new_guests = [(row["Nome"].strip(), email_registration)]
    guests_raw = row.get("Quem vem consigo?")

    if pd.notna(guests_raw):
        for guest in re.split(r"[;,]\s*", str(guests_raw)):
            guest = guest.strip()
            if not guest:
                continue
            match = re.match(r"(.+?)\s*\(([^)]+)\)", guest)
            name, email = match.groups() if match else (guest, email_registration)
            new_guests.append((name.strip(), email.strip()))

    for name, email in new_guests:
        complete_name = name
        short_name = extract_first_last(name)
        key = (str(submission_date), name.strip().lower(), email.strip().lower())
        if key not in existing_keys:
            output_rows.append({
                "date": submission_date,
                "email": email,
                "name": short_name,
                "complete_name": complete_name,
                "email_registration": email_registration,
                "supporting_document": supporting_document,
                "partition_date": today_str,
                "used": False,
                "sent": False
            })

# === Append new rows to CSV ===
if output_rows:
    new_df = pd.DataFrame(output_rows)
    final_df = pd.concat([existing_df, new_df], ignore_index=True)
    final_df.to_csv(GUEST_LIST_CSV_PATH, index=False)
    print(f"✅ {len(new_df)} novas linhas adicionadas a {GUEST_LIST_CSV_PATH}.")

    print("Nomes adicionados:")
    for row in output_rows:
        print(f"- {row['complete_name']} ({row['email_registration']})")

else:
    print("✅ Nenhuma nova linha para adicionar.")
